import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import random
import numpy as np
import os

# IPL 2025 Players with realistic stats
players_data = {
    'BATSMEN': [
        ('Virat Kohli', 'RCB', 'Both', 450, 1, 'Top Order', 'Wankhede Stadium'),
        ('Faf du Plessis', 'RCB', 'Pace', 380, 0, 'Top Order', '0'),
        ('Rajat Patidar', 'RCB', 'Spin', 320, 1, 'Middle Order', 'M. Chinnaswamy Stadium'),
        ('Will Jacks', 'RCB', 'Both', 280, 2, 'Top Order', '0'),
        ('Anuj Rawat', 'RCB', 'Spin', 200, 0, 'Middle Order', '0'),
        ('Suyash Prabhudessai', 'RCB', 'Both', 160, 1, 'Lower Middle Order', '0'),

        ('Ruturaj Gaikwad', 'CSK', 'Both', 420, 0, 'Top Order', 'M. A. Chidambaram Stadium'),
        ('Devon Conway', 'CSK', 'Spin', 390, 0, 'Top Order', 'M. A. Chidambaram Stadium'),
        ('Ajinkya Rahane', 'CSK', 'Spin', 300, 0, 'Middle Order', '0'),
        ('Shivam Dube', 'CSK', 'Pace', 260, 4, 'Lower Middle Order', 'Wankhede Stadium'),
        ('Rachin Ravindra', 'CSK', 'Both', 310, 2, 'Top Order', '0'),
        ('Sameer Rizvi', 'CSK', 'Both', 190, 0, 'Middle Order', '0'),

        ('Rohit Sharma', 'MI', 'Both', 440, 1, 'Top Order', 'Wankhede Stadium'),
        ('Ishan Kishan', 'MI', 'Pace', 370, 0, 'Top Order', 'Eden Gardens'),
        ('Suryakumar Yadav', 'MI', 'Both', 400, 0, 'Middle Order', 'Wankhede Stadium'),
        ('Tilak Varma', 'MI', 'Spin', 340, 1, 'Middle Order', '0'),
        ('Tim David', 'MI', 'Pace', 270, 2, 'Lower Middle Order', '0'),
        ('Dewald Brevis', 'MI', 'Pace', 220, 1, 'Middle Order', '0'),

        ('KL Rahul', 'LSG', 'Both', 410, 0, 'Top Order', 'Bharat Ratna Shri Atal Bihari Vajpayee Ekana Cricket Stadium'),
        ('Quinton de Kock', 'LSG', 'Pace', 390, 0, 'Top Order', '0'),
        ('Nicholas Pooran', 'LSG', 'Both', 310, 1, 'Lower Middle Order', '0'),
        ('Shubman Gill', 'GT', 'Both', 470, 0, 'Top Order', 'Narendra Modi Stadium'),
        ('David Warner', 'DC', 'Pace', 400, 1, 'Top Order', 'Arun Jaitley Stadium'),
        ('Prithvi Shaw', 'DC', 'Spin', 280, 0, 'Top Order', '0'),
        ('Jos Buttler', 'RR', 'Both', 430, 0, 'Top Order', 'Sawai Mansingh Stadium'),
        ('Sanju Samson', 'RR', 'Spin', 370, 1, 'Top Order', 'Sawai Mansingh Stadium'),
        ('Yashasvi Jaiswal', 'RR', 'Pace', 340, 0, 'Top Order', '0'),
        ('Mayank Agarwal', 'SRH', 'Both', 330, 0, 'Top Order', '0'),
        ('Aiden Markram', 'SRH', 'Spin', 320, 2, 'Top Order', '0'),
        ('Harry Brook', 'SRH', 'Pace', 360, 0, 'Middle Order', '0'),
        ('Shreyas Iyer', 'KKR', 'Spin', 350, 1, 'Top Order', 'Eden Gardens'),
        ('Rinku Singh', 'KKR', 'Pace', 240, 0, 'Lower Middle Order', 'Eden Gardens'),
        ('Phil Salt', 'KKR', 'Both', 330, 0, 'Top Order', '0'),
        ('Shahrukh Khan', 'PBKS', 'Pace', 200, 1, 'Lower Middle Order', '0'),
        ('Liam Livingstone', 'PBKS', 'Both', 290, 2, 'Middle Order', '0'),
    ],

    'BOWLERS': [
        ('Mohammed Siraj', 'RCB', 'Pace', 45, 18, 'Tailender', 'M. Chinnaswamy Stadium'),
        ('Josh Hazlewood', 'RCB', 'Pace', 35, 17, 'Tailender', '0'),
        ('Wanindu Hasaranga', 'RCB', 'Spin', 52, 21, 'Tailender', 'M. Chinnaswamy Stadium'),
        ('Akash Deep', 'RCB', 'Pace', 28, 12, 'Tailender', '0'),
        ('Karn Sharma', 'RCB', 'Spin', 38, 10, 'Tailender', '0'),
        ('Yash Dayal', 'RCB', 'Pace', 25, 13, 'Tailender', '0'),

        ('Deepak Chahar', 'CSK', 'Pace', 55, 16, 'Tailender', 'M. A. Chidambaram Stadium'),
        ('Tushar Deshpande', 'CSK', 'Pace', 42, 14, 'Tailender', '0'),
        ('Maheesh Theekshana', 'CSK', 'Spin', 38, 18, 'Tailender', '0'),
        ('Matheesha Pathirana', 'CSK', 'Pace', 30, 19, 'Tailender', 'M. A. Chidambaram Stadium'),
        ('Mukesh Choudhary', 'CSK', 'Pace', 40, 11, 'Tailender', '0'),
        ('Simarjeet Singh', 'CSK', 'Pace', 32, 12, 'Tailender', '0'),

        ('Jasprit Bumrah', 'MI', 'Pace', 38, 24, 'Tailender', 'Wankhede Stadium'),
        ('Piyush Chawla', 'MI', 'Spin', 45, 15, 'Tailender', '0'),
        ('Kumar Kartikeya', 'MI', 'Spin', 35, 12, 'Tailender', '0'),
        ('Jason Behrendorff', 'MI', 'Pace', 28, 16, 'Tailender', '0'),
        ('Akash Madhwal', 'MI', 'Pace', 22, 15, 'Tailender', '0'),
        ('Arjun Tendulkar', 'MI', 'Pace', 34, 9, 'Tailender', '0'),

        ('Rashid Khan', 'GT', 'Spin', 42, 22, 'Tailender', 'Narendra Modi Stadium'),
        ('Mohammed Shami', 'GT', 'Pace', 35, 20, 'Tailender', 'Narendra Modi Stadium'),
        ('Yuzvendra Chahal', 'RR', 'Spin', 48, 19, 'Tailender', 'Sawai Mansingh Stadium'),
        ('Trent Boult', 'RR', 'Pace', 40, 18, 'Tailender', '0'),
        ('Khaleel Ahmed', 'DC', 'Pace', 46, 14, 'Tailender', '0'),
        ('Kuldeep Yadav', 'DC', 'Spin', 42, 17, 'Tailender', 'Arun Jaitley Stadium'),
        ('Avesh Khan', 'LSG', 'Pace', 52, 16, 'Tailender', '0'),
        ('Ravi Bishnoi', 'LSG', 'Spin', 38, 16, 'Tailender', '0'),
        ('Arshdeep Singh', 'PBKS', 'Pace', 44, 18, 'Tailender', 'Punjab Cricket Association Stadium'),
        ('Rahul Chahar', 'PBKS', 'Spin', 35, 14, 'Tailender', '0'),
        ('Nathan Ellis', 'PBKS', 'Pace', 30, 15, 'Tailender', '0'),
        ('Lockie Ferguson', 'GT', 'Pace', 25, 19, 'Tailender', '0'),
        ('Noor Ahmad', 'GT', 'Spin', 38, 15, 'Tailender', '0'),
        ('Umran Malik', 'SRH', 'Pace', 32, 17, 'Tailender', '0'),
        ('T Natarajan', 'SRH', 'Pace', 42, 16, 'Tailender', 'Rajiv Gandhi International Stadium'),
        ('Bhuvneshwar Kumar', 'SRH', 'Pace', 48, 15, 'Tailender', '0'),
        ('Varun Chakaravarthy', 'KKR', 'Spin', 30, 18, 'Tailender', 'Eden Gardens'),
        ('Mitchell Starc', 'KKR', 'Pace', 38, 21, 'Tailender', '0'),
    ],

    'ALL_ROUNDERS': [
        ('Glenn Maxwell', 'RCB', 'Both', 270, 12, 'Middle Order', '0'),
        ('Mahipal Lomror', 'RCB', 'Spin', 200, 7, 'Lower Middle Order', '0'),
        ('Shahbaz Ahmed', 'RCB', 'Spin', 160, 8, 'Lower Middle Order', '0'),
        ('Cameron Green', 'RCB', 'Both', 250, 13, 'Middle Order', '0'),
        ('Ravindra Jadeja', 'CSK', 'Both', 230, 14, 'Lower Middle Order', 'M. A. Chidambaram Stadium'),
        ('Moeen Ali', 'CSK', 'Spin', 250, 10, 'Middle Order', '0'),
        ('Dwaine Pretorius', 'CSK', 'Pace', 150, 11, 'Lower Middle Order', '0'),
        ('Mitchell Santner', 'CSK', 'Spin', 180, 9, 'Lower Middle Order', '0'),
        ('Hardik Pandya', 'MI', 'Both', 290, 16, 'Lower Middle Order', 'Wankhede Stadium'),
        ('Hrithik Shokeen', 'MI', 'Spin', 120, 8, 'Lower Middle Order', '0'),
        ('Nehal Wadhera', 'MI', 'Both', 160, 4, 'Middle Order', '0'),
        ('Marcus Stoinis', 'LSG', 'Pace', 230, 12, 'Lower Middle Order', '0'),
        ('Krunal Pandya', 'LSG', 'Spin', 200, 10, 'Middle Order', '0'),
        ('Axar Patel', 'DC', 'Spin', 210, 13, 'Lower Middle Order', 'Arun Jaitley Stadium'),
        ('Lalit Yadav', 'DC', 'Both', 160, 7, 'Lower Middle Order', '0'),
        ('Rahul Tewatia', 'GT', 'Spin', 180, 9, 'Lower Middle Order', '0'),
        ('Sai Sudharsan', 'GT', 'Both', 210, 3, 'Middle Order', '0'),
        ('Riyan Parag', 'RR', 'Spin', 190, 8, 'Middle Order', '0'),
        ('Ravichandran Ashwin', 'RR', 'Spin', 150, 14, 'Tailender', '0'),
        ('Washington Sundar', 'SRH', 'Spin', 200, 11, 'Lower Middle Order', '0'),
        ('Abhishek Sharma', 'SRH', 'Both', 280, 9, 'Top Order', '0'),
        ('Marco Jansen', 'SRH', 'Pace', 130, 15, 'Tailender', '0'),
        ('Venkatesh Iyer', 'KKR', 'Pace', 220, 10, 'Top Order', '0'),
        ('Sunil Narine', 'KKR', 'Spin', 200, 17, 'Top Order', '0'),
        ('Andre Russell', 'KKR', 'Pace', 270, 18, 'Lower Middle Order', 'Eden Gardens'),
        ('Shardul Thakur', 'KKR', 'Both', 160, 12, 'Lower Middle Order', '0'),
        ('Nitish Rana', 'KKR', 'Spin', 250, 2, 'Middle Order', '0'),
        ('Shivam Mavi', 'LSG', 'Pace', 90, 10, 'Tailender', '0'),
        ('Deepak Hooda', 'LSG', 'Both', 220, 4, 'Middle Order', '0'),
        ('Rahul Tripathi', 'CSK', 'Both', 260, 1, 'Middle Order', '0'),
    ]
}

def generate_match_data(role, total_runs, total_wickets, batting_order):
    matches_data = []
    for i in range(10):
        if role == 'BATSMEN':
            if i < 3:
                runs = int(total_runs * random.uniform(0.12, 0.18))
            elif i < 7:
                runs = int(total_runs * random.uniform(0.06, 0.11))
            else:
                runs = int(total_runs * random.uniform(0.01, 0.05))
            wickets = 1 if (total_wickets > 0 and random.random() < 0.2) else 0
        elif role == 'BOWLERS':
            runs = random.randint(0, 6)
            if i < 4:
                wickets = max(0, int(total_wickets * random.uniform(0.12, 0.20)))
            elif i < 8:
                wickets = max(0, int(total_wickets * random.uniform(0.05, 0.11)))
            else:
                wickets = max(0, int(total_wickets * random.uniform(0.00, 0.04)))
        else:
            if batting_order in ['Top Order', 'Middle Order']:
                if i < 4:
                    runs = int(total_runs * random.uniform(0.10, 0.16))
                else:
                    runs = int(total_runs * random.uniform(0.03, 0.09))
            else:
                if i < 4:
                    runs = int(total_runs * random.uniform(0.08, 0.14))
                else:
                    runs = int(total_runs * random.uniform(0.02, 0.07))
            if i < 5:
                wickets = max(0, int(total_wickets * random.uniform(0.08, 0.18)))
            else:
                wickets = max(0, int(total_wickets * random.uniform(0.01, 0.08)))
        matches_data.append({'runs': max(0, runs), 'wickets': min(5, max(0, wickets))})
    return matches_data

wb = Workbook()
ws = wb.active
ws.title = "Players Data"

header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

headers = ['Player Name', 'Team', 'Skill (Spin/Pace/Both)', 'Total Runs (Last 10)',
           'Total Wickets (Last 10)', 'Batting Order', 'Favorite Stadium']
for i in range(1, 11):
    headers.extend([f'Match {i} Runs', f'Match {i} Wickets'])

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 50
ws.row_dimensions[1].height = 40

row = 2
for role, players in players_data.items():
    for player_info in players:
        name, team, skill, total_runs, total_wickets, batting_order, fav_stadium = player_info
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=2).value = team
        ws.cell(row=row, column=3).value = skill
        ws.cell(row=row, column=4).value = total_runs
        ws.cell(row=row, column=5).value = total_wickets
        ws.cell(row=row, column=6).value = batting_order
        ws.cell(row=row, column=7).value = fav_stadium
        matches = generate_match_data(role, total_runs, total_wickets, batting_order)
        col = 8
        for match in matches:
            ws.cell(row=row, column=col).value = match['runs']
            ws.cell(row=row, column=col+1).value = match['wickets']
            col += 2
        if row % 2 == 0:
            fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        row += 1

ws.freeze_panes = 'A2'

# Save to the data folder next to this script
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'players_data.xlsx')
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Players data saved to: {output_path}")
