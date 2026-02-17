from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

stadiums = [
    ('Wankhede Stadium', 'Mumbai', 'Red Soil', 'High', 'Both', 180, 6, 235, 263, 4, 172, 7, 224, 262, 5, 45, 'Y'),
    ('M. Chinnaswamy Stadium', 'Bengaluru', 'Red Soil', 'High', 'Pace', 185, 6, 263, 287, 3, 178, 7, 235, 262, 7, 38, 'Y'),
    ('M. A. Chidambaram Stadium', 'Chennai', 'Red Soil', 'Average', 'Spin', 168, 7, 246, 218, 5, 162, 8, 211, 203, 6, 52, 'N'),
    ('Eden Gardens', 'Kolkata', 'Clay Based', 'Average', 'Spin', 172, 7, 223, 272, 7, 165, 8, 206, 262, 2, 48, 'N'),
    ('Arun Jaitley Stadium', 'Delhi', 'Black Soil', 'High', 'Both', 178, 6, 231, 257, 4, 170, 7, 219, 246, 5, 42, 'N'),
    ('Rajiv Gandhi International Stadium', 'Hyderabad', 'Red Soil', 'Average', 'Pace', 170, 7, 227, 277, 3, 163, 8, 198, 253, 7, 50, 'N'),
    ('Sawai Mansingh Stadium', 'Jaipur', 'Black Soil', 'High', 'Pace', 178, 6, 226, 245, 4, 172, 7, 214, 232, 5, 40, 'N'),
    ('Punjab Cricket Association Stadium', 'Mohali', 'Red Soil', 'Average', 'Pace', 172, 7, 232, 248, 5, 166, 8, 207, 238, 6, 46, 'N'),
    ('Narendra Modi Stadium', 'Ahmedabad', 'Black Soil', 'Average', 'Both', 174, 7, 237, 242, 4, 168, 8, 203, 228, 6, 44, 'N'),
    ('Dr. Y.S. Rajasekhara Reddy ACA-VDCA Cricket Stadium', 'Visakhapatnam', 'Red Soil', 'Average', 'Pace', 167, 7, 195, 215, 6, 160, 8, 186, 198, 7, 54, 'N'),
    ('JSCA International Stadium', 'Ranchi', 'Red Soil', 'Low', 'Spin', 162, 8, 202, 208, 6, 155, 9, 188, 192, 7, 58, 'N'),
    ('Holkar Cricket Stadium', 'Indore', 'Black Soil', 'High', 'Pace', 182, 6, 235, 251, 3, 175, 7, 221, 243, 4, 36, 'N'),
    ('Barabati Stadium', 'Cuttack', 'Red Soil', 'Low', 'Spin', 160, 8, 214, 223, 6, 153, 9, 192, 207, 8, 56, 'N'),
    ('Vidarbha Cricket Association Stadium', 'Nagpur', 'Black Soil', 'Low', 'Spin', 158, 8, 209, 218, 6, 152, 9, 179, 196, 8, 60, 'N'),
    ('Maharashtra Cricket Association Stadium', 'Pune', 'Red Soil', 'Average', 'Both', 170, 7, 213, 231, 5, 164, 8, 197, 219, 7, 48, 'N'),
    ('Greenfield International Stadium', 'Thiruvananthapuram', 'Red Soil', 'Average', 'Spin', 164, 7, 199, 212, 6, 158, 8, 183, 195, 7, 52, 'N'),
    ('Himachal Pradesh Cricket Association Stadium', 'Dharamsala', 'Clay Based', 'Average', 'Pace', 160, 7, 218, 229, 5, 154, 8, 192, 211, 7, 50, 'N'),
    ('Bharat Ratna Shri Atal Bihari Vajpayee Ekana Cricket Stadium', 'Lucknow', 'Black Soil', 'Average', 'Both', 172, 7, 224, 257, 5, 166, 8, 201, 243, 6, 46, 'N'),
    ('Green Park Stadium', 'Kanpur', 'Black Soil', 'Low', 'Spin', 155, 8, 195, 204, 7, 148, 9, 171, 185, 8, 62, 'N'),
    ('VCA Stadium', 'Nagpur', 'Black Soil', 'Low', 'Spin', 157, 8, 206, 214, 6, 150, 9, 178, 193, 8, 58, 'N'),
    ('Saurashtra Cricket Association Stadium', 'Rajkot', 'Black Soil', 'High', 'Both', 176, 6, 227, 248, 4, 169, 7, 209, 235, 5, 40, 'N'),
    ('Shaheed Veer Narayan Singh International Stadium', 'Raipur', 'Red Soil', 'Average', 'Pace', 169, 7, 212, 226, 5, 162, 8, 195, 213, 7, 48, 'N'),
    ('Dr. DY Patil Sports Academy', 'Navi Mumbai', 'Red Soil', 'High', 'Both', 174, 6, 226, 241, 4, 168, 7, 207, 228, 5, 42, 'N'),
    ('Brabourne Stadium', 'Mumbai', 'Red Soil', 'High', 'Pace', 177, 6, 229, 244, 4, 171, 7, 213, 231, 5, 40, 'N'),
    ('IS Bindra Stadium', 'Mohali', 'Red Soil', 'Average', 'Pace', 171, 7, 230, 246, 5, 164, 8, 205, 233, 6, 46, 'N'),
]

wb = Workbook()
ws = wb.active
ws.title = "Stadiums Data"
header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
headers = ['Stadium Name', 'City', 'Pitch Type', 'Batting Scoring (High/Avg/Low)', 'Bowling Friendly (Pace/Spin/Both)',
    '1st Inn Avg Score', '1st Inn Avg Wickets', '1st Inn Highest Score', '1st Inn Highest Total', '1st Inn Highest Wickets',
    '2nd Inn Avg Score', '2nd Inn Avg Wickets', '2nd Inn Highest Score', '2nd Inn Highest Total', '2nd Inn Highest Wickets',
    'Total All-outs', 'Dew-Prone (Y/N)']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.column_dimensions['A'].width = 50
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
    ws.column_dimensions[col].width = 16
ws.row_dimensions[1].height = 50
for row, stadium in enumerate(stadiums, 2):
    for col, value in enumerate(stadium, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if row % 2 == 0:
            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        if col == 17 and value == 'Y':
            cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
            cell.font = Font(bold=True, color='FF4500')
ws.freeze_panes = 'A2'

# Save to the data folder next to this script
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'stadiums_data.xlsx')
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Stadiums data saved to: {output_path}")
